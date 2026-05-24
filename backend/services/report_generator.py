"""report_generator.py — PDF (WeasyPrint) and Excel (openpyxl) generation.

PDF matches the 5-page editorial design from report-pdf.jsx / report-pdf.css:
  Page 1 — Cover + KPI tiles + executive summary
  Page 2 — Task status breakdown + callouts
  Page 3 — Team leaderboard + champion callout + split stats
  Page 4 — Detailed task list with priority / status badges
  Page 5 — Daily throughput bar chart + methodology + colophon
"""
import html
import io
from datetime import datetime

# ── Design tokens ──────────────────────────────────────────────────────────────
_BG       = "#F6F3EC"
_BG_SOFT  = "#F0ECDF"
_SURFACE  = "#FCFAF5"
_RULE     = "#E2DCC9"
_RULE_SOFT= "#EFE9D8"
_INK      = "#1A2230"
_INK2     = "#4A5468"
_INK3     = "#6E7B91"
_INK_F    = "#A5B0C2"
_TEAL     = "#05b7aa"
_MID      = "#03a1b6"
_DEEP     = "#0082c6"
_OK       = "#0A7A6E"
_WARN     = "#B06A00"
_DANGER   = "#C0392B"

# pre-computed color-mix equivalents (WeasyPrint may not support color-mix)
_OK_BG    = "#EFF8F7"   # mix(ok 6%, surface)
_OK_BD    = "#ABCCC9"   # mix(ok 28%, rule)
_WARN_BG  = "#F7F2EC"
_WARN_BD  = "#CCBA9E"
_BAD_BG   = "#FAF4F4"
_BAD_BD   = "#D9CECA"
_CHAMP_BG = "#ECF8F7"   # mix(teal 10%, surface)
_CHAMP_BD = "#93D9D5"   # mix(teal 30%, rule)

_FONT_DISP  = 'Georgia, "Times New Roman", serif'
_FONT_UI    = '"Helvetica Neue", Arial, sans-serif'
_FONT_HINDI = '"Noto Serif Devanagari", serif'
_FONT_MONO  = '"Courier New", monospace'

_MEMBER_COLORS = ["#0082c6","#f59e0b","#05b7aa","#8b5cf6","#10b981","#ec4899","#ef4444","#6366f1"]


def _initials(name: str) -> str:
    parts = (name or "?").split()
    return "".join(p[0].upper() for p in parts[:2]) or "?"


def _fmt_mins(m: int) -> str:
    if not m: return "0h"
    h = m // 60; mn = m % 60
    return f"{h}h {mn}m" if mn else f"{h}h"


def _fmt_date(iso) -> str:
    if not iso: return "—"
    try:
        return datetime.fromisoformat(str(iso)).strftime("%-d %b")
    except Exception:
        return str(iso)[:10]


def _pri_color(pri: str) -> str:
    return {"urgent": "#dc2626", "high": "#ef4444", "medium": "#f59e0b", "low": "#10b981"}.get(
        (pri or "").lower(), "#6E7B91"
    )


def _status_style(status: str):
    """Return (bg, color, border) tuple for a status badge."""
    s = (status or "").lower()
    if s == "done":
        return "#D8EEEC", _OK, "#A3CEC9"
    if s in ("in_review", "review"):
        return "#EDE8FD", "#7c3aed", "#C4B8F8"
    if s == "in_progress":
        return "#DCF0FA", _DEEP, "#A8D4EF"
    return _BG_SOFT, _INK3, _RULE


def _page_shell(body_html: str, brand: str, meta_right: str, foot_left: str, page_n: int, page_of: int) -> str:
    return f"""
<div class="pdf">
  <div class="pdf__head">
    <div class="pdf__brand">
      <span class="pdf__brand-main">Kartavya</span>
      <span class="pdf__brand-hi">कर्तव्य</span>
      <span class="pdf__brand-by">by Aekam Inc</span>
    </div>
    <div class="pdf__head-meta">
      <span>{brand}</span>
      <span>{meta_right}</span>
    </div>
  </div>
  <div class="pdf__body">
    {body_html}
  </div>
  <div class="pdf__foot">
    <span>{foot_left}</span>
    <span>Page {page_n} / {page_of}</span>
  </div>
</div>"""


def _build_html(data: dict, team_name: str, period_from: str, period_to: str) -> str:
    team_name = html.escape(team_name)
    tasks          = data.get("tasks", {})
    entries        = data.get("entries", [])
    task_list      = data.get("task_list", [])
    by_member_t    = data.get("by_member_tasks", [])
    throughput     = data.get("daily_throughput", [])
    total_mins     = data.get("total_minutes", 0)

    todo        = tasks.get("todo", 0)
    in_progress = tasks.get("in_progress", 0)
    done        = tasks.get("done", 0)
    overdue     = tasks.get("overdue", 0)
    total_tasks = todo + in_progress + done

    total_h = _fmt_mins(total_mins)

    # per-member time map
    by_member_time: dict[str, int] = {}
    for e in entries:
        nm = e.get("user_name") or "Unknown"
        by_member_time[nm] = by_member_time.get(nm, 0) + (e.get("minutes") or 0)
    time_sorted = sorted(by_member_time.items(), key=lambda x: -x[1])

    generated = datetime.utcnow().strftime("%-d %b %Y · %H:%M UTC")
    period_label = f"{period_from} – {period_to}"

    # ── PAGE 1 — COVER ─────────────────────────────────────────────────────────
    # kpi tones
    done_kpi_cls  = "kpi--ok"
    over_kpi_cls  = "kpi--bad" if overdue > 0 else ""
    prog_kpi_cls  = "kpi--blue"

    # executive summary (auto-generated)
    champion_name = html.escape(by_member_t[0]["user_name"] if by_member_t else (time_sorted[0][0] if time_sorted else "the team"))
    exec_summary = (
        f"<b>{done} tasks completed</b> during this period across {len(time_sorted)} active members, "
        f"logging a combined {total_h} of tracked time. "
        f"{'There are ' + str(overdue) + ' overdue tasks requiring attention. ' if overdue > 0 else 'No tasks are currently overdue. '}"
        f"<b>{champion_name}</b> led the period with the most completed work. "
        f"{in_progress} tasks remain in progress heading into the next period."
    )

    cover_body = f"""
      <div class="pdf__cover">
        <div class="pdf__cover-kicker">PROJECT REPORT · {period_from.upper()}</div>
        <h1 class="pdf__cover-h1">
          {team_name}<br/>
          <em>{done} tasks done.</em>
        </h1>
        <div class="pdf__cover-hi">कार्य प्रतिवेदन</div>

        <div class="pdf__cover-meta">
          <div>
            <div class="pdf__meta-k">Project</div>
            <div class="pdf__meta-v">{team_name}</div>
          </div>
          <div>
            <div class="pdf__meta-k">Period</div>
            <div class="pdf__meta-v">{period_label}</div>
          </div>
          <div>
            <div class="pdf__meta-k">Members</div>
            <div class="pdf__meta-v">{len(time_sorted)} active</div>
          </div>
          <div>
            <div class="pdf__meta-k">Total time</div>
            <div class="pdf__meta-v">{total_h} logged</div>
          </div>
        </div>

        <div class="pdf__kpis">
          <div class="pdf__kpi pdf__kpi--ok">
            <div class="pdf__kpi-k">Completed</div>
            <div class="pdf__kpi-v">{done}</div>
            <div class="pdf__kpi-hint">tasks closed</div>
          </div>
          <div class="pdf__kpi pdf__kpi--blue">
            <div class="pdf__kpi-k">In Progress</div>
            <div class="pdf__kpi-v">{in_progress}</div>
            <div class="pdf__kpi-hint">active tasks</div>
          </div>
          <div class="pdf__kpi">
            <div class="pdf__kpi-k">To Do</div>
            <div class="pdf__kpi-v">{todo}</div>
            <div class="pdf__kpi-hint">not started</div>
          </div>
          <div class="pdf__kpi{'  pdf__kpi--bad' if overdue > 0 else ''}">
            <div class="pdf__kpi-k">Overdue</div>
            <div class="pdf__kpi-v">{overdue}</div>
            <div class="pdf__kpi-hint">past due date</div>
          </div>
        </div>

        <div class="pdf__exec">
          <div class="pdf__exec-h">Executive summary <span>सारांश</span></div>
          <p>{exec_summary}</p>
        </div>

        <div class="pdf__cover-cite">
          कर्तव्ये अधिकारस्ते मा फलेषु कदाचन। — <em>Bhagavad Gita 2.47</em>
        </div>
      </div>"""

    page1 = _page_shell(
        cover_body,
        f"Weekly report · {period_label}",
        "Cover",
        f"Generated {generated}",
        1, 5
    )

    # ── PAGE 2 — TASK BREAKDOWN ────────────────────────────────────────────────
    total_nz = max(total_tasks, 1)
    def status_bar(count, color):
        pct = round(count / total_nz * 100)
        return f'<div style="flex:1;height:8px;background:{_RULE_SOFT};border-radius:99px;overflow:hidden;"><div style="height:100%;width:{pct}%;background:{color};border-radius:99px;min-width:4px;"></div></div>'

    breakdown_rows = ""
    for label, count, color, hi in [
        ("Done",        done,        _OK,    "पूर्ण"),
        ("In Progress", in_progress, _DEEP,  "प्रगति में"),
        ("To Do",       todo,        _INK3,  "प्रतीक्षारत"),
        ("Overdue",     overdue,     _DANGER,"विलंबित"),
    ]:
        pct = round(count / total_nz * 100)
        bd_color = _DANGER if label == "Overdue" and count > 0 else _INK
        breakdown_rows += f"""
        <tr>
          <td style="padding:12px 10px;font-size:12px;font-weight:600;color:{_INK};vertical-align:middle;">{label} <span style="font-family:{_FONT_HINDI};font-size:11px;color:{_INK3};font-weight:400;">{hi}</span></td>
          <td style="padding:12px 10px;vertical-align:middle;">
            <div style="display:flex;align-items:center;gap:10px;">
              <div style="flex:1;height:8px;background:{_RULE_SOFT};border-radius:99px;overflow:hidden;">
                <div style="height:100%;width:{pct}%;background:{color};border-radius:99px;min-width:{'4' if count>0 else '0'}px;"></div>
              </div>
            </div>
          </td>
          <td style="padding:12px 10px;font-family:{_FONT_DISP};font-size:22px;font-weight:400;color:{color};text-align:right;vertical-align:middle;">{count}</td>
          <td style="padding:12px 10px;font-family:{_FONT_MONO};font-size:11px;color:{_INK3};text-align:right;vertical-align:middle;">{pct}%</td>
        </tr>"""

    # callouts
    best_callout = f"""
      <div class="pdf__callout">
        <div class="pdf__callout-k">Most completed</div>
        <div class="pdf__callout-v">{done} tasks done</div>
        <p>Completion rate {round(done/total_nz*100)}% of all tasks in scope.{' On track for the period.' if done >= in_progress else ''}</p>
      </div>"""
    warn_callout = ""
    if overdue > 0:
        warn_callout = f"""
      <div class="pdf__callout pdf__callout--warn">
        <div class="pdf__callout-k">Needs attention</div>
        <div class="pdf__callout-v">{overdue} overdue</div>
        <p>Tasks past their due date. Review and reassign or adjust timelines.</p>
      </div>"""
    else:
        warn_callout = f"""
      <div class="pdf__callout" style="border-left-color:{_OK};">
        <div class="pdf__callout-k">No overdue tasks</div>
        <div class="pdf__callout-v">All on track</div>
        <p>Zero tasks past their due date. Excellent execution this period.</p>
      </div>"""

    breakdown_body = f"""
      <div class="pdf__sec-h">
        <div>
          <h2>Task breakdown</h2>
          <span class="pdf__sec-hi">कार्य विश्लेषण</span>
        </div>
        <p>Status counts for all {total_tasks} tasks in <strong>{team_name}</strong> as of the report date.</p>
      </div>

      <table class="pdf__table" style="margin-bottom:18px;">
        <thead>
          <tr>
            <th>Status</th>
            <th>Distribution</th>
            <th class="num">Count</th>
            <th class="num">Share</th>
          </tr>
        </thead>
        <tbody>{breakdown_rows}</tbody>
      </table>

      <div class="pdf__callouts">
        {best_callout}
        {warn_callout}
      </div>

      <div style="margin-top:20px;">
        <div class="pdf__sec-h pdf__sec-h--tight">
          <div>
            <h2>Time logged</h2>
            <span class="pdf__sec-hi">समय</span>
          </div>
          <p>Time entries within the report period ({period_label}).</p>
        </div>
        <table class="pdf__table" style="margin-top:8px;">
          <thead>
            <tr>
              <th>Member</th>
              <th>Distribution</th>
              <th class="num">Hours</th>
              <th class="num">Share</th>
            </tr>
          </thead>
          <tbody>
            {''.join(
                f"""<tr>
                  <td style="font-size:12px;font-weight:500;">{html.escape(nm)}</td>
                  <td><div style="height:7px;background:{_RULE_SOFT};border-radius:99px;overflow:hidden;"><div style="height:100%;width:{round(mins/max(total_mins,1)*100)}%;background:linear-gradient(90deg,{_TEAL},{_DEEP});border-radius:99px;"></div></div></td>
                  <td class="num" style="font-size:14px;">{_fmt_mins(mins)}</td>
                  <td style="text-align:right;font-family:{_FONT_MONO};font-size:10px;color:{_INK3};">{round(mins/max(total_mins,1)*100)}%</td>
                </tr>"""
                for nm, mins in time_sorted[:8]
            ) if time_sorted else f'<tr><td colspan="4" style="padding:14px 10px;color:{_INK3};">No time entries for this period.</td></tr>'}
          </tbody>
        </table>
      </div>"""

    page2 = _page_shell(
        breakdown_body,
        f"{team_name} · {period_label}",
        "Task breakdown",
        "कालः सृजति भूतानि — Time creates all things.",
        2, 5
    )

    # ── PAGE 3 — LEADERBOARD ───────────────────────────────────────────────────
    # merge by_member_tasks and by_member_time into unified leaderboard
    all_names = list({nm for nm, _ in time_sorted} | {r["user_name"] for r in by_member_t})
    tasks_map = {r["user_name"]: r["tasks_done"] for r in by_member_t}
    leaderboard = sorted(
        [(nm, tasks_map.get(nm, 0), by_member_time.get(nm, 0)) for nm in all_names],
        key=lambda x: (-x[1], -x[2])
    )
    max_tasks = max((x[1] for x in leaderboard), default=1) or 1

    champion = leaderboard[0] if leaderboard else ("—", 0, 0)
    champ_name_esc = html.escape(champion[0])
    champ_initials = _initials(champion[0])
    champ_color = _MEMBER_COLORS[0]

    champion_block = f"""
      <div class="pdf__champ">
        <div class="pdf__champ-l">CHAMPION OF THE PERIOD <span>अवधि के सर्वश्रेष्ठ</span></div>
        <div class="pdf__champ-row">
          <div class="pdf__champ-av" style="background:{champ_color};">{champ_initials}</div>
          <div>
            <div class="pdf__champ-name">{champ_name_esc}</div>
            <div class="pdf__champ-role">{team_name}</div>
          </div>
          <div class="pdf__champ-stats">
            <div><b>{champion[1]}</b><span>tasks done</span></div>
            <div><b>{_fmt_mins(champion[2])}</b><span>time logged</span></div>
          </div>
        </div>
        <p class="pdf__champ-note">
          Led the team with {champion[1]} task{'s' if champion[1] != 1 else ''} completed
          and {_fmt_mins(champion[2])} of tracked time during {period_label}.
        </p>
      </div>"""

    board_rows = ""
    for i, (nm, tc, mins) in enumerate(leaderboard[:8]):
        nm_esc = html.escape(nm)
        color = _MEMBER_COLORS[i % len(_MEMBER_COLORS)]
        bar_pct = round(tc / max_tasks * 100) if tc else 0
        board_rows += f"""
        <div class="pdf__board-row">
          <span class="pdf__board-rank">{i+1}</span>
          <span class="pdf__board-av" style="background:{color};">{_initials(nm)}</span>
          <div class="pdf__board-id">
            <div class="pdf__board-name">{nm_esc}</div>
            <div class="pdf__board-role">{_fmt_mins(mins)} logged</div>
          </div>
          <div class="pdf__board-bar"><div style="width:{bar_pct}%;background:{color};"></div></div>
          <span class="pdf__board-n">{tc}</span>
        </div>"""

    avg_tasks = round(sum(x[1] for x in leaderboard) / max(len(leaderboard), 1), 1)
    split_cards = f"""
      <div class="pdf__split">
        <div class="pdf__split-card">
          <div class="pdf__split-h">Total time logged</div>
          <div class="pdf__split-n">{total_h}<small>this period</small></div>
          <p>Across {len(time_sorted)} member{'s' if len(time_sorted) != 1 else ''}.
          Average <b>{_fmt_mins(total_mins // max(len(time_sorted), 1))}</b> per member.</p>
        </div>
        <div class="pdf__split-card">
          <div class="pdf__split-h">Tasks closed</div>
          <div class="pdf__split-n">{done}<small>completed</small></div>
          <p>Average <b>{avg_tasks}</b> tasks per active member.
          {f'<b>{overdue}</b> still overdue.' if overdue else 'No overdue tasks.'}</p>
        </div>
      </div>"""

    leaderboard_body = f"""
      {champion_block if leaderboard else ''}
      <div class="pdf__sec-h pdf__sec-h--tight">
        <div>
          <h2>Leaderboard</h2>
          <span class="pdf__sec-hi">वरीयता क्रम</span>
        </div>
        <p>Ranked by tasks completed in the period. Time logged as tie-breaker.</p>
      </div>
      <div class="pdf__board">
        {board_rows if board_rows else f'<p style="color:{_INK3};font-size:12px;padding:12px 0;">No member data for this period.</p>'}
      </div>
      {split_cards}"""

    page3 = _page_shell(
        leaderboard_body,
        f"{team_name} · {period_label}",
        "Team performance",
        "कर्मण्येवाधिकारस्ते — You have a right only to action.",
        3, 5
    )

    # ── PAGE 4 — TASK LIST ─────────────────────────────────────────────────────
    task_rows = ""
    for i, t in enumerate(task_list[:40]):
        pri   = (t.get("priority") or "medium").lower()
        pri_color = _pri_color(pri)
        st    = t.get("status") or "todo"
        st_bg, st_color, st_bd = _status_style(st)
        st_label = st.replace("_", " ").title()
        due_str  = _fmt_date(t.get("due_at"))
        owner    = html.escape((t.get("owner_name") or "—")[:18])
        title    = html.escape((t.get("title") or "Untitled")[:55])
        short_id = f"#{str(t.get('task_id', ''))[-5:]}"
        task_rows += f"""
        <tr>
          <td class="mono">{short_id}</td>
          <td class="ttl">{title}</td>
          <td><span class="pdf__pri" style="background:{pri_color};">{pri.capitalize()}</span></td>
          <td class="mono">{owner}</td>
          <td>{due_str}</td>
          <td><span class="pdf__st" style="background:{st_bg};color:{st_color};border-color:{st_bd};">{st_label}</span></td>
        </tr>"""

    task_list_body = f"""
      <div class="pdf__sec-h">
        <div>
          <h2>Detailed task list</h2>
          <span class="pdf__sec-hi">कार्य सूची</span>
        </div>
        <p>All tasks in {team_name} ({len(task_list)} total · showing first {min(len(task_list),40)}).</p>
      </div>

      <table class="pdf__tasks">
        <thead>
          <tr>
            <th style="width:54px;">ID</th>
            <th>Title</th>
            <th style="width:62px;">Priority</th>
            <th style="width:90px;">Owner</th>
            <th style="width:52px;">Due</th>
            <th style="width:82px;">Status</th>
          </tr>
        </thead>
        <tbody>
          {task_rows if task_rows else f'<tr><td colspan="6" style="padding:16px 8px;color:{_INK3};">No tasks found.</td></tr>'}
        </tbody>
      </table>

      <div class="pdf__legend">
        <span><i style="background:#dc2626;"></i>Urgent</span>
        <span><i style="background:#ef4444;"></i>High</span>
        <span><i style="background:#f59e0b;"></i>Medium</span>
        <span><i style="background:#10b981;"></i>Low</span>
        <span class="pdf__legend-sep">·</span>
        <span>ID shows last 5 chars of task identifier</span>
      </div>"""

    page4 = _page_shell(
        task_list_body,
        f"{team_name} · {period_label}",
        "Detailed task list",
        "प्रयत्नशील् — Keep striving.",
        4, 5
    )

    # ── PAGE 5 — THROUGHPUT + METHODOLOGY ─────────────────────────────────────
    max_daily = max((r["done_count"] for r in throughput), default=1) or 1
    chart_cols = ""
    if throughput:
        for r in throughput:
            day_label = r["day"][5:]  # MM-DD
            cnt = r["done_count"]
            bar_h = max(round(cnt / max_daily * 120), 4)
            chart_cols += f"""
            <div class="pdf__chart-col">
              <div class="pdf__chart-stack">
                <div class="pdf__chart-done" style="height:{bar_h}px;"></div>
              </div>
              <div class="pdf__chart-lbl">{day_label}</div>
              <div class="pdf__chart-n">{cnt}</div>
            </div>"""
    else:
        chart_cols = f'<p style="color:{_INK3};font-size:12px;margin:auto;">No tasks closed during this period.</p>'

    trend_body = f"""
      <div class="pdf__sec-h">
        <div>
          <h2>Throughput trend</h2>
          <span class="pdf__sec-hi">गति</span>
        </div>
        <p>Tasks closed each day in the report period.</p>
      </div>

      <div class="pdf__chart">
        {chart_cols}
      </div>
      <div class="pdf__chart-key">
        <span><i class="k1"></i>Tasks completed per day</span>
      </div>

      <div class="pdf__sec-h pdf__sec-h--tight">
        <div>
          <h2>Methodology &amp; data</h2>
          <span class="pdf__sec-hi">विधि</span>
        </div>
      </div>

      <div class="pdf__method">
        <div>
          <div class="pdf__method-k">Source</div>
          <p>All data is pulled directly from the Kartavya production database (Postgres). No third-party aggregator.</p>
        </div>
        <div>
          <div class="pdf__method-k">Counting rules</div>
          <p><b>Completed</b> = tasks in Done status. <b>Overdue</b> = unfinished tasks past their due date. <b>Time</b> = sum of all time entries started within the period.</p>
        </div>
        <div>
          <div class="pdf__method-k">Champion</div>
          <p>Member with the most tasks completed in the period. Ties broken by total time logged.</p>
        </div>
        <div>
          <div class="pdf__method-k">Period</div>
          <p>Report covers <b>{period_from}</b> to <b>{period_to}</b> (inclusive). Generated {generated}.</p>
        </div>
      </div>

      <div class="pdf__colophon">
        <div class="pdf__colophon-h">
          <span class="pdf__colophon-main">Kartavya</span>
          <span class="pdf__colophon-hi">कर्तव्य</span>
          <em>— do what must be done.</em>
        </div>
        <p>
          Aekam Inc · Project: <b>{team_name}</b> ·
          Report period {period_label} · Generated {generated}.
          This report is confidential to admins and project owners.
        </p>
        <p class="pdf__colophon-cite">
          कर्म एव अधिकारस्ते — <em>your right is to your work alone.</em>
        </p>
      </div>"""

    page5 = _page_shell(
        trend_body,
        f"{team_name} · {period_label}",
        "Throughput & methodology",
        "कर्मण्येवाधिकारस्ते — You have a right only to action.",
        5, 5
    )

    # ── CSS ────────────────────────────────────────────────────────────────────
    css = f"""
/* system fonts only — no external requests from WeasyPrint */

*{{ box-sizing:border-box; margin:0; padding:0; }}
body{{ background:{_BG}; font-family:{_FONT_UI}; color:{_INK}; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
@page{{ size:A4; margin:0; }}

.pdf{{
  width:210mm; height:297mm;
  background:{_SURFACE};
  color:{_INK};
  display:flex; flex-direction:column;
  padding:48px 56px 36px;
  position:relative;
  page-break-after:always;
  font-family:{_FONT_UI};
}}
.pdf::before{{
  content:""; position:absolute; left:0; right:0; top:0; height:6px;
  background:linear-gradient(90deg,{_TEAL},{_MID},{_DEEP});
}}

/* Head */
.pdf__head{{
  display:flex; align-items:flex-start; justify-content:space-between; gap:20px;
  padding-bottom:12px; border-bottom:1px solid {_RULE_SOFT}; margin-bottom:20px;
}}
.pdf__brand{{ display:flex; align-items:baseline; gap:8px; }}
.pdf__brand-main{{ font-family:{_FONT_DISP}; font-size:17px; font-weight:500; color:{_INK}; letter-spacing:-0.005em; }}
.pdf__brand-hi{{ font-family:{_FONT_HINDI}; font-size:13px; color:{_MID}; }}
.pdf__brand-by{{ font-size:9px; letter-spacing:0.22em; text-transform:uppercase; color:{_INK3}; font-weight:700; margin-left:6px; }}
.pdf__head-meta{{ text-align:right; font-size:10px; color:{_INK3}; line-height:1.6; }}
.pdf__head-meta span:first-child{{ display:block; font-weight:600; color:{_INK2}; letter-spacing:0.04em; }}

/* Body */
.pdf__body{{ flex:1; min-height:0; display:flex; flex-direction:column; gap:16px; overflow:hidden; }}

/* Foot */
.pdf__foot{{
  margin-top:14px; padding-top:10px; border-top:1px solid {_RULE_SOFT};
  display:flex; justify-content:space-between;
  font-size:9px; color:{_INK_F}; font-family:{_FONT_MONO}; letter-spacing:0.02em;
}}

/* Cover */
.pdf__cover{{ display:flex; flex-direction:column; gap:16px; }}
.pdf__cover-kicker{{ font-size:10px; letter-spacing:0.28em; text-transform:uppercase; color:{_MID}; font-weight:700; }}
.pdf__cover-h1{{
  font-family:{_FONT_DISP}; font-size:52px; font-weight:400; line-height:1.02;
  letter-spacing:-0.025em; margin:0; color:{_INK};
}}
.pdf__cover-h1 em{{ font-family:{_FONT_DISP}; font-style:italic; color:{_DEEP}; font-weight:400; }}
.pdf__cover-hi{{ font-family:{_FONT_HINDI}; font-size:20px; color:{_TEAL}; margin-top:-4px; }}
.pdf__cover-meta{{
  display:grid; grid-template-columns:repeat(4,1fr); gap:14px;
  padding:12px 0 4px; border-top:1px dashed {_RULE}; border-bottom:1px dashed {_RULE}; margin-top:4px;
}}
.pdf__meta-k{{ font-size:9px; letter-spacing:0.2em; text-transform:uppercase; color:{_INK3}; font-weight:700; margin-bottom:3px; }}
.pdf__meta-v{{ font-family:{_FONT_DISP}; font-size:14px; color:{_INK}; font-weight:500; }}
.pdf__kpis{{ display:grid; grid-template-columns:repeat(4,1fr); gap:8px; margin-top:4px; }}
.pdf__kpi{{
  background:{_BG_SOFT}; border:1px solid {_RULE}; border-radius:9px; padding:12px 12px 10px;
  display:flex; flex-direction:column; gap:3px;
}}
.pdf__kpi-k{{ font-size:8px; letter-spacing:0.18em; text-transform:uppercase; color:{_INK3}; font-weight:700; }}
.pdf__kpi-v{{ font-family:{_FONT_DISP}; font-size:36px; font-weight:400; line-height:1; letter-spacing:-0.025em; color:{_INK}; margin-top:2px; }}
.pdf__kpi-hint{{ font-size:10px; color:{_INK3}; margin-top:auto; }}
.pdf__kpi--ok{{ border-color:{_OK_BD}; background:{_OK_BG}; }}
.pdf__kpi--ok .pdf__kpi-v{{ color:{_OK}; }}
.pdf__kpi--blue{{ border-color:#A8D4EF; background:#DCF0FA; }}
.pdf__kpi--blue .pdf__kpi-v{{ color:{_DEEP}; }}
.pdf__kpi--bad{{ border-color:{_BAD_BD}; background:{_BAD_BG}; }}
.pdf__kpi--bad .pdf__kpi-v{{ color:{_DANGER}; }}
.pdf__exec{{ margin-top:6px; }}
.pdf__exec-h{{
  font-family:{_FONT_DISP}; font-size:16px; font-weight:500; color:{_INK}; margin-bottom:6px;
  display:flex; align-items:baseline; gap:10px;
}}
.pdf__exec-h span{{ font-family:{_FONT_HINDI}; font-size:13px; color:{_INK3}; font-weight:400; }}
.pdf__exec p{{ margin:0 0 6px; font-size:12.5px; line-height:1.6; color:{_INK2}; }}
.pdf__exec b{{ color:{_INK}; font-weight:600; }}
.pdf__cover-cite{{
  margin-top:auto; padding-top:12px; border-top:1px solid {_RULE_SOFT};
  font-family:{_FONT_HINDI}; font-size:12px; color:{_INK2};
}}
.pdf__cover-cite em{{ font-family:{_FONT_DISP}; font-style:italic; color:{_INK3}; }}

/* Section header */
.pdf__sec-h{{
  display:flex; align-items:flex-end; justify-content:space-between; gap:20px;
  padding-bottom:8px; border-bottom:1px solid {_RULE}; margin-bottom:6px;
}}
.pdf__sec-h--tight{{ margin-top:8px; }}
.pdf__sec-h h2{{
  margin:0; font-family:{_FONT_DISP}; font-size:26px; font-weight:400; color:{_INK};
  letter-spacing:-0.015em; line-height:1.05;
}}
.pdf__sec-hi{{ font-family:{_FONT_HINDI}; font-size:14px; color:{_MID}; margin-left:8px; }}
.pdf__sec-h p{{ max-width:300px; margin:0 0 3px; font-size:10px; color:{_INK3}; line-height:1.5; text-align:right; }}

/* Tables */
.pdf__table{{ width:100%; border-collapse:collapse; font-size:12px; }}
.pdf__table th{{
  text-align:left; font-size:8.5px; letter-spacing:0.16em; text-transform:uppercase;
  color:{_INK3}; font-weight:700; padding:9px 10px; border-bottom:1px solid {_RULE};
}}
.pdf__table th.num{{ text-align:right; }}
.pdf__table td{{ padding:10px 10px; border-bottom:1px dashed {_RULE}; color:{_INK}; vertical-align:middle; }}
.pdf__table td.num{{
  text-align:right; font-family:{_FONT_DISP}; font-size:18px; font-weight:400; letter-spacing:-0.01em;
}}

/* Callouts */
.pdf__callouts{{ display:grid; grid-template-columns:1fr 1fr; gap:10px; margin-top:6px; }}
.pdf__callout{{
  background:{_BG_SOFT}; border:1px solid {_RULE}; border-left:3px solid {_OK}; border-radius:8px; padding:12px 14px;
}}
.pdf__callout--warn{{ border-left-color:{_WARN}; }}
.pdf__callout-k{{ font-size:8.5px; letter-spacing:0.2em; text-transform:uppercase; color:{_INK3}; font-weight:700; }}
.pdf__callout-v{{ font-family:{_FONT_DISP}; font-size:17px; color:{_INK}; font-weight:500; margin:4px 0 5px; }}
.pdf__callout p{{ margin:0; font-size:11px; color:{_INK2}; line-height:1.55; }}

/* Champion */
.pdf__champ{{
  background:{_CHAMP_BG}; border:1px solid {_CHAMP_BD}; border-radius:11px; padding:16px 20px;
}}
.pdf__champ-l{{ font-size:9px; letter-spacing:0.26em; text-transform:uppercase; color:{_MID}; font-weight:700; margin-bottom:12px; }}
.pdf__champ-l span{{ font-family:{_FONT_HINDI}; font-size:11px; color:{_INK3}; font-weight:400; letter-spacing:0; text-transform:none; margin-left:8px; }}
.pdf__champ-row{{ display:flex; align-items:center; gap:14px; }}
.pdf__champ-av{{
  width:56px; height:56px; border-radius:50%; color:#fff; font-weight:700; font-size:20px;
  display:flex; align-items:center; justify-content:center; flex-shrink:0;
}}
.pdf__champ-name{{ font-family:{_FONT_DISP}; font-size:26px; font-weight:500; color:{_INK}; letter-spacing:-0.01em; }}
.pdf__champ-role{{ font-size:12px; color:{_INK3}; margin-top:2px; }}
.pdf__champ-stats{{ margin-left:auto; display:flex; gap:20px; }}
.pdf__champ-stats > div{{ text-align:center; }}
.pdf__champ-stats b{{ display:block; font-family:{_FONT_DISP}; font-size:22px; font-weight:400; color:{_INK}; letter-spacing:-0.02em; line-height:1; }}
.pdf__champ-stats span{{ font-size:8.5px; letter-spacing:0.14em; text-transform:uppercase; color:{_INK3}; font-weight:700; margin-top:3px; display:block; }}
.pdf__champ-note{{ margin:12px 0 0; font-size:12px; color:{_INK2}; line-height:1.6; font-style:italic; font-family:{_FONT_DISP}; }}

/* Leaderboard */
.pdf__board{{ display:flex; flex-direction:column; }}
.pdf__board-row{{
  display:flex; align-items:center; gap:12px;
  padding:9px 0; border-bottom:1px dashed {_RULE};
}}
.pdf__board-row:last-child{{ border-bottom:0; }}
.pdf__board-rank{{ font-family:{_FONT_DISP}; font-size:17px; color:{_INK3}; width:22px; text-align:center; flex-shrink:0; }}
.pdf__board-av{{ width:28px; height:28px; border-radius:50%; color:#fff; font-weight:700; font-size:10px; display:flex; align-items:center; justify-content:center; flex-shrink:0; }}
.pdf__board-id{{ flex:0 0 160px; }}
.pdf__board-name{{ font-size:12.5px; font-weight:600; color:{_INK}; line-height:1.2; }}
.pdf__board-role{{ font-size:10px; color:{_INK3}; margin-top:1px; }}
.pdf__board-bar{{ flex:1; height:7px; background:{_RULE_SOFT}; border-radius:99px; overflow:hidden; }}
.pdf__board-bar > div{{ height:100%; border-radius:99px; min-width:5px; }}
.pdf__board-n{{ text-align:right; font-family:{_FONT_DISP}; font-size:17px; color:{_INK}; width:28px; flex-shrink:0; }}

/* Split cards */
.pdf__split{{ display:grid; grid-template-columns:1fr 1fr; gap:12px; margin-top:8px; }}
.pdf__split-card{{ background:{_BG_SOFT}; border:1px solid {_RULE}; border-radius:9px; padding:12px 14px; }}
.pdf__split-h{{ font-size:8.5px; letter-spacing:0.2em; text-transform:uppercase; color:{_INK3}; font-weight:700; }}
.pdf__split-n{{
  font-family:{_FONT_DISP}; font-size:30px; font-weight:400; color:{_INK}; letter-spacing:-0.02em;
  line-height:1; margin:5px 0 7px; display:flex; align-items:baseline; gap:7px;
}}
.pdf__split-n small{{ font-family:{_FONT_UI}; font-size:9px; letter-spacing:0.16em; text-transform:uppercase; color:{_INK3}; font-weight:700; }}
.pdf__split-card p{{ margin:0; font-size:11px; color:{_INK2}; line-height:1.55; }}
.pdf__split-card b{{ color:{_INK}; font-weight:600; }}

/* Task table */
.pdf__tasks{{ width:100%; border-collapse:collapse; font-size:11px; table-layout:fixed; }}
.pdf__tasks th{{ text-align:left; font-size:8.5px; letter-spacing:0.15em; text-transform:uppercase; color:{_INK3}; font-weight:700; padding:8px 6px; border-bottom:1px solid {_RULE}; }}
.pdf__tasks td{{ padding:8px 6px; border-bottom:1px dashed {_RULE}; color:{_INK}; vertical-align:middle; }}
.pdf__tasks td.mono{{ font-family:{_FONT_MONO}; font-size:10px; color:{_INK2}; }}
.pdf__tasks td.ttl{{ font-weight:500; }}

.pdf__pri{{
  display:inline-block; font-size:8.5px; letter-spacing:0.08em; text-transform:uppercase;
  font-weight:700; padding:2px 6px; border-radius:99px; color:#fff;
}}
.pdf__st{{
  display:inline-block; font-size:8.5px; letter-spacing:0.06em; text-transform:uppercase;
  font-weight:700; padding:2px 7px; border-radius:99px; border:1px solid;
}}

/* Legend */
.pdf__legend{{ display:flex; flex-wrap:wrap; gap:10px; margin-top:6px; font-size:10px; color:{_INK3}; align-items:center; }}
.pdf__legend span{{ display:inline-flex; align-items:center; gap:5px; }}
.pdf__legend i{{ width:8px; height:8px; border-radius:50%; display:inline-block; }}
.pdf__legend-sep{{ color:{_RULE}; }}
.pdf__legend b{{ color:{_INK2}; font-weight:600; }}

/* Chart */
.pdf__chart{{
  background:{_BG_SOFT}; border:1px solid {_RULE}; border-radius:9px;
  padding:20px 20px 14px; display:flex; align-items:flex-end; gap:12px; height:200px;
}}
.pdf__chart-col{{ flex:1; display:flex; flex-direction:column; align-items:center; gap:5px; height:100%; justify-content:flex-end; }}
.pdf__chart-stack{{ width:100%; max-width:40px; display:flex; flex-direction:column-reverse; border-radius:4px 4px 0 0; overflow:hidden; }}
.pdf__chart-done{{ width:100%; background:linear-gradient(180deg,{_TEAL},{_DEEP}); }}
.pdf__chart-lbl{{ font-size:8.5px; letter-spacing:0.1em; text-transform:uppercase; color:{_INK3}; font-weight:700; }}
.pdf__chart-n{{ font-family:{_FONT_DISP}; font-size:12px; color:{_INK}; }}
.pdf__chart-key{{ display:flex; gap:14px; margin-top:5px; font-size:10px; color:{_INK3}; }}
.pdf__chart-key span{{ display:flex; align-items:center; gap:5px; }}
.pdf__chart-key i{{ width:12px; height:7px; border-radius:2px; }}
.pdf__chart-key i.k1{{ background:linear-gradient(90deg,{_TEAL},{_DEEP}); }}

/* Methodology */
.pdf__method{{ display:grid; grid-template-columns:1fr 1fr; gap:12px 20px; margin-top:4px; }}
.pdf__method-k{{ font-size:8.5px; letter-spacing:0.2em; text-transform:uppercase; color:{_INK3}; font-weight:700; margin-bottom:3px; }}
.pdf__method p{{ margin:0; font-size:11px; color:{_INK2}; line-height:1.55; }}
.pdf__method b{{ color:{_INK}; font-weight:600; }}

/* Colophon */
.pdf__colophon{{ margin-top:auto; padding-top:14px; border-top:1px solid {_RULE}; }}
.pdf__colophon-h{{ display:flex; align-items:baseline; gap:9px; margin-bottom:7px; }}
.pdf__colophon-main{{ font-family:{_FONT_DISP}; font-size:17px; font-weight:500; color:{_INK}; }}
.pdf__colophon-hi{{ font-family:{_FONT_HINDI}; font-size:13px; color:{_MID}; }}
.pdf__colophon-h em{{ font-family:{_FONT_DISP}; font-style:italic; color:{_INK3}; font-size:11px; }}
.pdf__colophon p{{ margin:0 0 5px; font-size:10.5px; color:{_INK3}; line-height:1.55; }}
.pdf__colophon b{{ color:{_INK2}; font-weight:600; }}
.pdf__colophon-cite{{ font-family:{_FONT_HINDI}; font-size:12px; color:{_INK2}; margin-top:7px; }}
.pdf__colophon-cite em{{ font-family:{_FONT_DISP}; font-style:italic; color:{_INK3}; font-size:10.5px; }}

/* Proj breakdown */
.pdf__proj{{ display:flex; align-items:center; gap:8px; }}
.pdf__proj > i{{ width:7px; height:7px; border-radius:2px; flex-shrink:0; }}
.pdf__proj-name{{ font-size:12px; font-weight:600; color:{_INK}; line-height:1.2; }}
.pdf__proj-hi{{ font-family:{_FONT_HINDI}; font-size:10.5px; color:{_INK3}; margin-top:1px; }}
"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<style>
{css}
</style>
</head>
<body>
{page1}
{page2}
{page3}
{page4}
{page5}
</body>
</html>"""


def generate_pdf(data: dict, team_name: str, period_from: str, period_to: str) -> bytes:
    try:
        from weasyprint import HTML
    except ImportError as e:
        raise RuntimeError("WeasyPrint is not available on this server") from e
    html = _build_html(data, team_name, period_from, period_to)
    return HTML(string=html, base_url=None).write_pdf()


def generate_excel(data: dict, team_name: str, period_from: str, period_to: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tasks      = data.get("tasks", {})
    entries    = data.get("entries", [])
    task_list  = data.get("task_list", [])
    by_mt      = data.get("by_member_tasks", [])

    C_INK  = "FF1A2230"; C_INK3 = "FF6E7B91"; C_TEAL = "FF05b7aa"
    C_DEEP = "FF0082c6"; C_BG   = "FFF6F3EC"; C_SURF = "FFFCFAF5"; C_HDR = "FF1A2230"

    thin = Border(bottom=Side(style="thin", color="FFE2DCC9"))

    wb = openpyxl.Workbook()

    def hdr_row(ws, row, *values):
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=C_HDR)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin

    def data_row(ws, row, *values, bold=False, color=None):
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(name="Calibri", bold=bold, size=10,
                          color=color or (C_DEEP if bold else C_INK))
            c.fill = PatternFill("solid", fgColor=C_SURF if row % 2 == 0 else C_BG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin

    total_mins = data.get("total_minutes", 0)
    total_h = _fmt_mins(total_mins)

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 28; ws1.column_dimensions["B"].width = 18
    ws1["A1"] = f"Kartavya — {team_name} Report"
    ws1["A1"].font = Font(name="Georgia", bold=True, size=16, color=C_INK)
    ws1["A2"] = f"Period: {period_from} – {period_to}"
    ws1["A2"].font = Font(name="Calibri", size=10, color=C_INK3)
    ws1["A3"] = f"Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}"
    ws1["A3"].font = Font(name="Calibri", size=10, italic=True, color=C_INK3)
    hdr_row(ws1, 5, "Metric", "Value")
    summary_data = [
        ("Total Time Logged", total_h),
        ("Total Tasks",       tasks.get("todo", 0) + tasks.get("in_progress", 0) + tasks.get("done", 0)),
        ("Done",              tasks.get("done", 0)),
        ("In Progress",       tasks.get("in_progress", 0)),
        ("To Do",             tasks.get("todo", 0)),
        ("Overdue",           tasks.get("overdue", 0)),
        ("Time Entries",      len(entries)),
    ]
    for i, (label, val) in enumerate(summary_data, 6):
        data_row(ws1, i, label, val)

    # ── Sheet 2: Tasks ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Tasks")
    for col, w in zip("ABCDEFG", [9, 40, 14, 12, 18, 12, 10]):
        ws2.column_dimensions[col].width = w
    hdr_row(ws2, 1, "ID", "Title", "Status", "Priority", "Owner", "Due", "Updated")
    for i, t in enumerate(task_list, 2):
        short_id = str(t.get("task_id", ""))[-8:]
        due_str = str(t.get("due_at", "") or "")[:10]
        upd_str = str(t.get("updated_at", "") or "")[:10]
        data_row(ws2, i,
                 short_id,
                 (t.get("title") or "")[:100],
                 (t.get("status") or "").replace("_", " ").title(),
                 (t.get("priority") or "").capitalize(),
                 t.get("owner_name") or "",
                 due_str, upd_str)

    # ── Sheet 3: Time Entries ─────────────────────────────────────────
    ws3 = wb.create_sheet("Time Entries")
    for col, w in zip("ABCDE", [12, 22, 40, 36, 12]):
        ws3.column_dimensions[col].width = w
    hdr_row(ws3, 1, "Date", "Member", "Task", "Note", "Hours")
    for i, e in enumerate(entries, 2):
        date_str = ""
        if e.get("started_at"):
            try: date_str = datetime.fromisoformat(str(e["started_at"])).strftime("%d %b %Y")
            except Exception: date_str = str(e["started_at"])[:10]
        data_row(ws3, i, date_str,
                 e.get("user_name") or "",
                 (e.get("task_title") or "")[:100],
                 (e.get("description") or "")[:120],
                 round((e.get("minutes") or 0) / 60, 2))

    # ── Sheet 4: By Member ────────────────────────────────────────────
    ws4 = wb.create_sheet("By Member")
    for col, w in zip("ABCD", [26, 14, 14, 14]):
        ws4.column_dimensions[col].width = w
    hdr_row(ws4, 1, "Member", "Tasks Done", "Time (h)", "% Time Share")
    by_member_time: dict[str, int] = {}
    for e in entries:
        nm = e.get("user_name") or "Unknown"
        by_member_time[nm] = by_member_time.get(nm, 0) + (e.get("minutes") or 0)
    tasks_map = {r["user_name"]: r["tasks_done"] for r in by_mt}
    all_members = sorted(set(list(by_member_time.keys()) + list(tasks_map.keys())))
    for i, nm in enumerate(sorted(all_members,
                                   key=lambda x: (-tasks_map.get(x,0), -by_member_time.get(x,0))), 2):
        mins = by_member_time.get(nm, 0)
        pct = round(mins / max(total_mins, 1) * 100, 1)
        data_row(ws4, i, nm, tasks_map.get(nm, 0), round(mins / 60, 2), f"{pct}%", bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
